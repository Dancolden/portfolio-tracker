#!/usr/bin/env python3
"""
Portfolio Tracker Updater
=========================
Run this script once a month (or whenever you update the spreadsheet) to
regenerate portfolio_tracker.html with fresh data.

Usage:
    python update_tracker.py

Requirements:
    pip install openpyxl

Both files must be in the same folder:
    update_tracker.py
    portfolio_tracker.html
    Live_Updating_Investment_Club_Portfolio_Tracker.xlsx  (or edit XLSX_FILE below)
"""

import json
import math
import os
import sys
import re
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
XLSX_FILE = "Live Updating Investment Club Portfolio Tracker.xlsx"
HTML_FILE = "index.html"
# ─────────────────────────────────────────────────────────────────────────────


def find_file(filename):
    """Look for file in script's directory."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, filename)
    if os.path.exists(path):
        return path
    # Also check current working directory
    if os.path.exists(filename):
        return os.path.abspath(filename)
    return None


def load_spreadsheet(xlsx_path):
    print(f"  Loading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return wb


def extract_daily_returns(wb):
    """
    Extract daily returns from Greek Scratch Work sheet.
    col C = daily portfolio simple return
    col E = daily SPX simple return
    col B = month-end date label (EOMONTH)
    Matches the spreadsheet's SUMIF-based monthly aggregation exactly.
    """
    ws = wb['Greek Scratch Work']
    rows = list(ws.iter_rows(values_only=True))

    daily_data = []
    for r in rows[1:]:
        date, month_label, port_ret, spx_ret = r[0], r[1], r[2], r[4]
        if (isinstance(port_ret, float) and
                isinstance(spx_ret, float) and
                hasattr(month_label, 'year')):
            daily_data.append({
                'date': date.strftime('%Y-%m-%d'),
                'ym':   f"{month_label.year}-{month_label.month:02d}",
                'p':    round(port_ret, 10),
                's':    round(spx_ret,  10),
            })

    print(f"  Daily returns: {len(daily_data)} rows "
          f"({daily_data[0]['date']} → {daily_data[-1]['date']})")
    return daily_data


def extract_perf_data(wb):
    """
    Extract normalized performance series from Equity Price Data sheet.
    Normalizes portfolio + all indexes to base 100 at inception.
    """
    ws = wb['Equity Price Data']
    rows = list(ws.iter_rows(values_only=True))

    raw = []
    for row in rows[2:]:
        if not row[0]:
            continue
        try:
            raw.append({
                'date':      row[0],
                'portfolio': float(row[1]) if row[1] else None,
                'spx':       float(row[2]) if row[2] else None,
                'nasdaq':    float(row[3]) if row[3] else None,
                'dji':       float(row[4]) if row[4] else None,
                'rut':       float(row[5]) if row[5] else None,
            })
        except (TypeError, ValueError):
            continue

    # Find first row where all series have data
    base = next((r for r in raw if all(r[k] for k in ['portfolio','spx','nasdaq','dji','rut'])), None)
    if not base:
        raise ValueError("Could not find a base row with all series populated")

    normalized = []
    for r in raw:
        rec = {'date': r['date'].strftime('%Y-%m-%d')}
        for k in ['portfolio', 'spx', 'nasdaq', 'dji', 'rut']:
            rec[k] = round(r[k] / base[k] * 100, 4) if r[k] and base[k] else None
        normalized.append(rec)

    print(f"  Performance data: {len(normalized)} rows "
          f"({normalized[0]['date']} → {normalized[-1]['date']})")
    return normalized



def extract_transactions(wb):
    """
    Extract full transaction log from Equity Volume Data sheet.
    Detects share count changes (new positions + buys/sells).
    Cross-references Equity Price Data for prices.
    """
    # Build price lookup
    ws_price = wb['Equity Price Data']
    price_rows = list(ws_price.iter_rows(values_only=True))
    price_headers = list(price_rows[0])

    price_lookup = {}
    for row in price_rows[2:]:
        if not row[0]: continue
        date_str = row[0].strftime('%Y-%m-%d')
        price_lookup[date_str] = {}
        for i, ticker in enumerate(price_headers[1:], start=1):
            if ticker and row[i] and isinstance(row[i], (int, float)):
                price_lookup[date_str][ticker] = row[i]

    def get_price(ticker, date_str):
        for d in reversed(sorted(d for d in price_lookup if d <= date_str)):
            if ticker in price_lookup[d]:
                return price_lookup[d][ticker]
        return None

    ws = wb['Equity Volume Data']
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    tickers = headers[5:]

    prev_shares = {}
    transactions = []

    for row in rows[1:]:
        date = row[0]
        if not hasattr(date, 'strftime'): continue
        date_str = date.strftime('%Y-%m-%d')
        for col_i, ticker in enumerate(tickers, start=5):
            if ticker is None or ticker == 'CASH': continue
            shares = row[col_i]
            if not isinstance(shares, (int, float)): continue
            prev = prev_shares.get(ticker)
            if prev is None and shares > 0:
                price = get_price(ticker, date_str)
                transactions.append({'date': date_str, 'action': 'BUY', 'ticker': ticker,
                                      'change': shares, 'prev': 0, 'new': shares, 'price': price})
            elif prev is not None and abs(shares - prev) > 0.01:
                change = shares - prev
                price = get_price(ticker, date_str)
                transactions.append({'date': date_str, 'action': 'BUY' if change > 0 else 'SELL',
                                      'ticker': ticker, 'change': abs(change), 'prev': prev,
                                      'new': shares, 'price': price})
            prev_shares[ticker] = shares

    transactions.sort(key=lambda x: x['date'], reverse=True)
    print(f"  Transactions: {len(transactions)} ({sum(1 for t in transactions if t['action']=='BUY')} buys, "
          f"{sum(1 for t in transactions if t['action']=='SELL')} sells)")
    return transactions


def verify_metrics(daily_data):
    """Compute and print the risk metrics so you can cross-check."""
    DAILY_RF = 0.05 / 252

    # Monthly returns via SUMIF method
    month_p, month_s = defaultdict(float), defaultdict(float)
    for d in daily_data:
        month_p[d['ym']] += d['p']
        month_s[d['ym']] += d['s']

    months  = sorted(month_p.keys())
    port_m  = [month_p[m] for m in months]
    spx_m   = [month_s[m] for m in months]
    n       = len(port_m)

    pm = sum(port_m) / n
    sm = sum(spx_m)  / n
    cov  = sum((port_m[i]-pm)*(spx_m[i]-sm) for i in range(n)) / (n-1)
    var_s= sum((spx_m[i]-sm)**2              for i in range(n)) / (n-1)
    beta  = cov / var_s
    alpha = pm - beta * sm

    daily_p = [d['p'] for d in daily_data]
    dp      = sum(daily_p) / len(daily_p)
    std_p   = math.sqrt(sum((r-dp)**2 for r in daily_p) / len(daily_p))
    sharpe  = (dp - DAILY_RF) / std_p * math.sqrt(252)

    print(f"\n  Computed metrics ({n} monthly obs, {len(daily_p)} daily obs):")
    print(f"    Beta   = {beta:.4f}")
    print(f"    Alpha  = {alpha*100:.4f}%/month  ({alpha*12*100:.2f}% annualised)")
    print(f"    Sharpe = {sharpe:.4f}")
    return beta, alpha, sharpe


def inject_into_html(html_path, daily_data, perf_data, txn_data=None):
    """Replace DAILY_RETURNS and PERF_DATA constants in the HTML file."""
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    daily_json = json.dumps(daily_data, separators=(',', ':'))
    perf_json  = json.dumps(perf_data,  separators=(',', ':'))

    # Replace DAILY_RETURNS
    pattern_daily = r'const DAILY_RETURNS\s*=\s*\[.*?\];'
    new_daily     = f'const DAILY_RETURNS = {daily_json};'
    if re.search(pattern_daily, content, re.DOTALL):
        content = re.sub(pattern_daily, new_daily, content, flags=re.DOTALL)
        print("  ✓ DAILY_RETURNS updated")
    else:
        print("  ✗ DAILY_RETURNS not found in HTML — was the file modified?")
        return False

    # Replace PERF_DATA
    pattern_perf = r'const PERF_DATA\s*=\s*\[.*?\];'
    new_perf     = f'const PERF_DATA = {perf_json};'
    if re.search(pattern_perf, content, re.DOTALL):
        content = re.sub(pattern_perf, new_perf, content, flags=re.DOTALL)
        print("  ✓ PERF_DATA updated")
    else:
        print("  ✗ PERF_DATA not found in HTML — was the file modified?")
        return False

    # Replace TRANSACTION_DATA
    if txn_data is not None:
        txn_json = json.dumps(txn_data, separators=(',', ':'))
        pattern_txn = r'const TRANSACTION_DATA\s*=\s*\[.*?\];'
        new_txn     = f'const TRANSACTION_DATA = {txn_json};'
        if re.search(pattern_txn, content, re.DOTALL):
            content = re.sub(pattern_txn, new_txn, content, flags=re.DOTALL)
            print("  ✓ TRANSACTION_DATA updated")
        else:
            print("  ⚠ TRANSACTION_DATA not found — skipping")

    # Stamp the update date in the HTML for reference
    today = datetime.today().strftime('%Y-%m-%d')
    content = re.sub(
        r'<!-- last-updated: .*? -->',
        f'<!-- last-updated: {today} -->',
        content
    )
    if '<!-- last-updated:' not in content:
        content = content.replace('</head>', f'<!-- last-updated: {today} -->\n</head>', 1)

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(content)

    size_kb = os.path.getsize(html_path) / 1024
    print(f"  ✓ HTML saved ({size_kb:.0f} KB)")
    return True


def main():
    print("=" * 55)
    print("  Portfolio Tracker Updater")
    print(f"  {datetime.today().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 55)

    # Locate files
    xlsx_path = find_file(XLSX_FILE)
    html_path = find_file(HTML_FILE)

    if not xlsx_path:
        print(f"\nERROR: Cannot find '{XLSX_FILE}'")
        print(f"  Place the spreadsheet in the same folder as this script.")
        sys.exit(1)
    if not html_path:
        print(f"\nERROR: Cannot find '{HTML_FILE}'")
        print(f"  Place portfolio_tracker.html in the same folder as this script.")
        sys.exit(1)

    print(f"\n[1/4] Loading spreadsheet...")
    wb = load_spreadsheet(xlsx_path)

    print(f"\n[2/4] Extracting daily returns (for Beta/Alpha/Sharpe)...")
    daily_data = extract_daily_returns(wb)

    print(f"\n[3/4] Extracting performance series (for charts)...")
    perf_data = extract_perf_data(wb)

    print(f"\n[4/4] Extracting transaction log...")
    txn_data = extract_transactions(wb)

    print(f"\n[5/5] Injecting data into HTML...")
    ok = inject_into_html(html_path, daily_data, perf_data, txn_data)

    if ok:
        verify_metrics(daily_data)
        print("\n" + "=" * 55)
        print("  Done! Next steps:")
        print("  1. Open index.html to verify it looks correct")
        print("  2. git add index.html")
        print("  3. git commit -m \'Update data\'")
        print("  4. git push")
        print("=" * 55)
    else:
        print("\nERROR: HTML injection failed. No changes were saved.")
        sys.exit(1)


if __name__ == '__main__':
    main()
